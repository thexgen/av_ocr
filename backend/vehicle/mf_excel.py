"""Detect and parse Asset Vantage Mutual Fund Excel templates (.xls / .xlsx)."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

_HEADER_ALIASES: dict[str, str] = {
    "mf_name": "mf_name",
    "investor_name": "investor_name",
    "entity_id": "entity_id",
    "pan_number": "pan_number",
    "folio_number": "folio_number",
    "currency": "currency",
    "product_code": "product_code",
    "isin": "isin",
    "symbol": "symbol",
    "cusip": "cusip",
    "scheme_name": "scheme_name",
    "trade_date": "trade_date",
    "transaction_type": "transaction_type",
    "dividend_rate": "dividend_rate",
    "amount": "amount",
    "transfer_date": "transfer_date",
    "transfer_price": "transfer_price",
    "transfer_amount": "transfer_amount",
    "units": "units",
    "price": "price",
    "advisor": "advisor",
    "arn_code": "arn_code",
    "brokerage": "brokerage",
    "stt_charges": "stt_charges",
    "service_tax": "service_tax",
    "transaction_charges": "transaction_charges",
    "stamp_duty": "stamp_duty",
    "other_charges": "other_charges",
    "tds": "tds",
    "notes": "notes",
    "transaction_id": "transaction_id",
    "transaction_qualifier": "transaction_qualifier",
    "specific_lot": "specific_lot",
}

_MF_MARKERS = {
    "folio_number",
    "scheme_name",
    "mf_name",
    "units",
    "trade_date",
    "transaction_type",
}


def _norm_header(raw: Any) -> str:
    text = str(raw or "").strip().lower()
    text = text.replace("*", " ")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _to_decimal(raw: Any) -> Decimal | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, Decimal):
        return raw
    if isinstance(raw, (int, float)):
        return Decimal(str(raw))
    text = str(raw).strip().replace(",", "")
    if not text:
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _to_date(raw: Any) -> date | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, float):
        # Excel serial date via xlrd
        try:
            import xlrd  # type: ignore

            return datetime(*xlrd.xldate_as_tuple(raw, 0)[:3]).date()
        except Exception:  # noqa: BLE001
            return None
    text = str(raw).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y", "%d %b %Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _to_int(raw: Any) -> int | None:
    if raw is None or raw == "":
        return None
    try:
        return int(float(str(raw).strip()))
    except (TypeError, ValueError):
        return None


def _read_sheet_matrix(content: bytes, filename: str) -> tuple[str, list[list[Any]]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xls":
        import xlrd  # type: ignore

        book = xlrd.open_workbook(file_contents=content)
        sheet = book.sheet_by_index(0)
        rows = [
            [sheet.cell_value(r, c) for c in range(sheet.ncols)]
            for r in range(sheet.nrows)
        ]
        return sheet.name, rows

    # .xlsx and unknown excel-like → openpyxl
    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    name = ws.title
    wb.close()
    return name, rows


def detect_vehicle_type(
    *,
    filename: str,
    sheet_name: str,
    headers: list[str],
) -> str | None:
    """Return mutual-fund | fixed-income | direct-equity | None."""
    name_l = filename.lower()
    sheet_l = sheet_name.lower().replace(" ", "").replace("_", "")
    header_set = set(headers)

    if (
        "mutualfund" in name_l
        or "mutual_fund" in name_l
        or "mutualfund" in sheet_l
        or ({"folio_number", "scheme_name"} <= header_set)
        or ({"mf_name", "units", "trade_date"} <= header_set)
    ):
        return "mutual-fund"

    if (
        "fixedincome" in name_l
        or "fixed_income" in name_l
        or "fixedincome" in sheet_l
        or (
            "security_name" in header_set
            and "isin" in header_set
            and "transaction_type" in header_set
        )
    ):
        return "fixed-income"

    if (
        "directequity" in name_l
        or "direct_equity" in name_l
        or "directequity" in sheet_l
        or ({"scrip_name", "isin"} <= header_set)
    ):
        return "direct-equity"

    # Soft MF marker score
    hits = len(header_set & _MF_MARKERS)
    if hits >= 4:
        return "mutual-fund"
    return None


@dataclass
class MfParsedRow:
    entity_id: int | None = None
    investor_name: str | None = None
    mf_name: str | None = None
    folio_number: str | None = None
    isin: str | None = None
    scheme_name: str | None = None
    trade_date: date | None = None
    transaction_type: str | None = None
    dividend_rate: Decimal | None = None
    amount: Decimal | None = None
    units: Decimal | None = None
    price: Decimal | None = None
    transfer_date: date | None = None
    transfer_price: Decimal | None = None
    transfer_amount: Decimal | None = None
    brokerage: Decimal | None = None
    stt_charges: Decimal | None = None
    service_tax: Decimal | None = None
    transaction_charges: Decimal | None = None
    stamp_duty: Decimal | None = None
    other_charges: Decimal | None = None
    tds: Decimal | None = None
    notes: str | None = None
    transaction_id: str | None = None
    transaction_qualifier: str | None = None
    currency: str | None = None
    iserror: int = 0
    errordesc: str | None = None


@dataclass
class MfParseResult:
    vehicle_type: str
    sheet_name: str
    headers: list[str]
    rows: list[MfParsedRow] = field(default_factory=list)
    raw_row_count: int = 0


def parse_mutual_fund_excel(content: bytes, filename: str) -> MfParseResult:
    sheet_name, matrix = _read_sheet_matrix(content, filename)
    if not matrix:
        raise ValueError("Excel file is empty.")

    header_cells = matrix[0]
    mapped_headers: list[str | None] = []
    for cell in header_cells:
        key = _norm_header(cell)
        mapped_headers.append(_HEADER_ALIASES.get(key))

    canonical_headers = [h for h in mapped_headers if h]
    vehicle = detect_vehicle_type(
        filename=filename,
        sheet_name=sheet_name,
        headers=canonical_headers,
    )
    if vehicle != "mutual-fund":
        raise ValueError(
            f"File does not look like a Mutual Fund template "
            f"(detected={vehicle or 'unknown'}, sheet={sheet_name!r})."
        )

    parsed_rows: list[MfParsedRow] = []
    for raw in matrix[1:]:
        if not raw or all(v is None or str(v).strip() == "" for v in raw):
            continue
        values: dict[str, Any] = {}
        for idx, field_name in enumerate(mapped_headers):
            if not field_name or idx >= len(raw):
                continue
            values[field_name] = raw[idx]

        row = MfParsedRow(
            entity_id=_to_int(values.get("entity_id")),
            investor_name=_str(values.get("investor_name")),
            mf_name=_str(values.get("mf_name")),
            folio_number=_str(values.get("folio_number")),
            isin=_str(values.get("isin")),
            scheme_name=_str(values.get("scheme_name")),
            trade_date=_to_date(values.get("trade_date")),
            transaction_type=_str(values.get("transaction_type")),
            dividend_rate=_to_decimal(values.get("dividend_rate")),
            amount=_to_decimal(values.get("amount")),
            units=_to_decimal(values.get("units")),
            price=_to_decimal(values.get("price")),
            transfer_date=_to_date(values.get("transfer_date")),
            transfer_price=_to_decimal(values.get("transfer_price")),
            transfer_amount=_to_decimal(values.get("transfer_amount")),
            brokerage=_to_decimal(values.get("brokerage")),
            stt_charges=_to_decimal(values.get("stt_charges")),
            service_tax=_to_decimal(values.get("service_tax")),
            transaction_charges=_to_decimal(values.get("transaction_charges")),
            stamp_duty=_to_decimal(values.get("stamp_duty")),
            other_charges=_to_decimal(values.get("other_charges")),
            tds=_to_decimal(values.get("tds")),
            notes=_str(values.get("notes")),
            transaction_id=_str(values.get("transaction_id")),
            transaction_qualifier=_str(values.get("transaction_qualifier")),
            currency=_str(values.get("currency")),
        )
        errors: list[str] = []
        if not row.trade_date:
            errors.append("TRADE_DATE missing")
        if not row.transaction_type:
            errors.append("TRANSACTION_TYPE missing")
        if row.amount is None and row.units is None:
            errors.append("AMOUNT/UNITS missing")
        if not row.folio_number and not row.scheme_name and not row.isin:
            errors.append("FOLIO/SCHEME/ISIN missing")
        if errors:
            row.iserror = 1
            row.errordesc = "; ".join(errors)
        parsed_rows.append(row)

    return MfParseResult(
        vehicle_type="mutual-fund",
        sheet_name=sheet_name,
        headers=canonical_headers,
        rows=parsed_rows,
        raw_row_count=max(0, len(matrix) - 1),
    )


def peek_vehicle_type(content: bytes, filename: str) -> tuple[str | None, str, list[str]]:
    """Lightweight detect without full row parse."""
    sheet_name, matrix = _read_sheet_matrix(content, filename)
    if not matrix:
        return None, sheet_name, []
    headers = []
    for cell in matrix[0]:
        key = _norm_header(cell)
        mapped = _HEADER_ALIASES.get(key)
        if mapped:
            headers.append(mapped)
        else:
            # keep normalized unknown headers for FI/DE detection
            if key:
                headers.append(key)
    vehicle = detect_vehicle_type(
        filename=filename, sheet_name=sheet_name, headers=headers
    )
    return vehicle, sheet_name, headers


def _str(raw: Any) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    return text or None
