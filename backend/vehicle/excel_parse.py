"""Parse AV Mutual Fund / Fixed Income / Direct Equity Excel templates."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

# Normalized header -> canonical field
_HEADER_ALIASES: dict[str, str] = {
    # shared / MF
    "mf_name": "mf_name",
    "investor_name": "investor_name",
    "entity_id": "entity_id",
    "entity_name": "entity_name",
    "entity_name_first_holder": "entity_name",
    "pan_number": "pan_number",
    "folio_number": "folio_number",
    "currency": "currency",
    "product_code": "product_code",
    "isin": "isin",
    "symbol": "symbol",
    "cusip": "cusip",
    "scheme_name": "scheme_name",
    "trade_date": "trade_date",
    "transaction_date": "trade_date",
    "transaction_type": "transaction_type",
    "dividend_rate": "dividend_rate",
    "amount": "amount",
    "net_amount": "amount",
    "final_transaction_amount": "amount",
    "transfer_date": "transfer_date",
    "transfer_price": "transfer_price",
    "transfer_amount": "transfer_amount",
    "units": "units",
    "quantity": "units",
    "price": "price",
    "clean_price": "price",
    "dirty_price": "price",
    "advisor": "advisor",
    "arn_code": "arn_code",
    "brokerage": "brokerage",
    "stt_charges": "stt_charges",
    "service_tax": "service_tax",
    "transaction_charges": "transaction_charges",
    "stamp_duty": "stamp_duty",
    "other_charges": "other_charges",
    "tds": "tds",
    "tds_amount": "tds",
    "notes": "notes",
    "transaction_id": "transaction_id",
    "transaction_qualifier": "transaction_qualifier",
    "specific_lot": "specific_lot",
    # FI / DE
    "security_name": "security_name",
    "scrip_name": "security_name",
    "exchange": "exchange",
    "position_type": "position_type",
    "depository_custodian_name": "custodian",
    "depository_custodian_number": "custodian_no",
    "dp_client_id_no_custodian_account_number": "custodian_no",
}


def _norm_header(raw: Any) -> str:
    text = str(raw or "").strip().lower()
    text = text.replace("*", " ")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _to_decimal(raw: Any) -> Decimal | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, Decimal):
        return raw
    if isinstance(raw, (int, float)):
        return Decimal(str(raw))
    text = str(raw).strip().replace(",", "")
    if not text:
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _to_date(raw: Any) -> date | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, float):
        try:
            import xlrd  # type: ignore

            return datetime(*xlrd.xldate_as_tuple(raw, 0)[:3]).date()
        except Exception:  # noqa: BLE001
            return None
    text = str(raw).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y", "%d %b %Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _to_int(raw: Any) -> int | None:
    if raw is None or raw == "":
        return None
    try:
        return int(float(str(raw).strip()))
    except (TypeError, ValueError):
        return None


def _str(raw: Any) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    return text or None


def read_sheet_matrix(content: bytes, filename: str) -> tuple[str, list[list[Any]]]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".xls":
        import xlrd  # type: ignore

        book = xlrd.open_workbook(file_contents=content)
        sheet = book.sheet_by_index(0)
        rows = [
            [sheet.cell_value(r, c) for c in range(sheet.ncols)]
            for r in range(sheet.nrows)
        ]
        return sheet.name, rows

    from openpyxl import load_workbook

    wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    name = ws.title
    wb.close()
    return name, rows


def map_headers(header_cells: list[Any]) -> list[str | None]:
    mapped: list[str | None] = []
    for cell in header_cells:
        key = _norm_header(cell)
        mapped.append(_HEADER_ALIASES.get(key))
    return mapped


def detect_vehicle_type(
    *,
    filename: str,
    sheet_name: str,
    headers: list[str],
) -> str | None:
    name_l = filename.lower()
    sheet_l = sheet_name.lower().replace(" ", "").replace("_", "")
    header_set = set(headers)

    if (
        "mutualfund" in name_l
        or "mutual_fund" in name_l
        or "mutualfund" in sheet_l
        or ({"folio_number", "scheme_name"} <= header_set)
        or ({"mf_name", "units", "trade_date"} <= header_set)
    ):
        return "mutual-fund"

    if (
        "fixedincome" in name_l
        or "fixed_income" in name_l
        or "fixedincome" in sheet_l
        or (
            "security_name" in header_set
            and "isin" in header_set
            and "transaction_type" in header_set
            and "scrip_name" not in header_set
            and "folio_number" not in header_set
        )
    ):
        # Prefer DE if scrip markers stronger — checked below first for DE filename
        if "directequity" not in name_l and "direct_equity" not in name_l:
            return "fixed-income"

    if (
        "directequity" in name_l
        or "direct_equity" in name_l
        or "directequity" in sheet_l
        or ({"security_name", "isin", "exchange"} <= header_set and "folio_number" not in header_set)
    ):
        # DE templates use Scrip Name → security_name; FI uses Security Name too.
        # Distinguish: DE often has position_type / exchange + no accrued interest header.
        if "fixedincome" in sheet_l or "fixed_income" in name_l:
            return "fixed-income"
        return "direct-equity"

    if "security_name" in header_set and "isin" in header_set:
        return "fixed-income"

    hits = len(header_set & {"folio_number", "scheme_name", "mf_name", "units", "trade_date"})
    if hits >= 4:
        return "mutual-fund"
    return None


@dataclass
class ParsedVehicleRow:
    entity_id: int | None = None
    trade_date: date | None = None
    transaction_type: str | None = None
    isin: str | None = None
    security_name: str | None = None
    folio_number: str | None = None
    scheme_name: str | None = None
    mf_name: str | None = None
    units: Decimal | None = None
    price: Decimal | None = None
    amount: Decimal | None = None
    dividend_rate: Decimal | None = None
    brokerage: Decimal | None = None
    stt_charges: Decimal | None = None
    service_tax: Decimal | None = None
    transaction_charges: Decimal | None = None
    stamp_duty: Decimal | None = None
    other_charges: Decimal | None = None
    tds: Decimal | None = None
    notes: str | None = None
    transaction_id: str | None = None
    transaction_qualifier: str | None = None
    transfer_date: date | None = None
    transfer_price: Decimal | None = None
    transfer_amount: Decimal | None = None
    iserror: int = 0
    errordesc: str | None = None


@dataclass
class VehicleParseResult:
    vehicle_type: str
    sheet_name: str
    headers: list[str]
    rows: list[ParsedVehicleRow] = field(default_factory=list)


def peek_vehicle_type(content: bytes, filename: str) -> tuple[str | None, str, list[str]]:
    sheet_name, matrix = read_sheet_matrix(content, filename)
    if not matrix:
        return None, sheet_name, []
    mapped = map_headers(matrix[0])
    headers = [h for h in mapped if h]
    # also keep raw norms for detection edge cases
    for cell in matrix[0]:
        key = _norm_header(cell)
        if key and key not in headers and key not in _HEADER_ALIASES:
            headers.append(key)
    vehicle = detect_vehicle_type(
        filename=filename, sheet_name=sheet_name, headers=[h for h in mapped if h]
    )
    # retry with extended headers if needed
    if vehicle is None:
        vehicle = detect_vehicle_type(
            filename=filename, sheet_name=sheet_name, headers=headers
        )
    return vehicle, sheet_name, [h for h in mapped if h]


def parse_vehicle_excel(
    content: bytes,
    filename: str,
    *,
    expect: str | None = None,
) -> VehicleParseResult:
    sheet_name, matrix = read_sheet_matrix(content, filename)
    if not matrix:
        raise ValueError("Excel file is empty.")

    mapped_headers = map_headers(matrix[0])
    canonical = [h for h in mapped_headers if h]
    vehicle = detect_vehicle_type(
        filename=filename, sheet_name=sheet_name, headers=canonical
    )
    if expect and vehicle and vehicle != expect:
        # still allow if filename/sheet strongly match expect
        pass
    if expect and not vehicle:
        vehicle = expect
    if not vehicle:
        raise ValueError(
            f"Could not identify Excel template (sheet={sheet_name!r})."
        )
    if expect and vehicle != expect:
        raise ValueError(
            f"Expected {expect} file but detected {vehicle} (sheet={sheet_name!r})."
        )

    parsed_rows: list[ParsedVehicleRow] = []
    for raw in matrix[1:]:
        if not raw or all(v is None or str(v).strip() == "" for v in raw):
            continue
        values: dict[str, Any] = {}
        for idx, field_name in enumerate(mapped_headers):
            if not field_name or idx >= len(raw):
                continue
            values[field_name] = raw[idx]

        security = _str(values.get("security_name")) or _str(values.get("scheme_name")) or _str(
            values.get("mf_name")
        )
        row = ParsedVehicleRow(
            entity_id=_to_int(values.get("entity_id")),
            trade_date=_to_date(values.get("trade_date")),
            transaction_type=_str(values.get("transaction_type")),
            isin=_str(values.get("isin")),
            security_name=security,
            folio_number=_str(values.get("folio_number")),
            scheme_name=_str(values.get("scheme_name")),
            mf_name=_str(values.get("mf_name")),
            units=_to_decimal(values.get("units")),
            price=_to_decimal(values.get("price")),
            amount=_to_decimal(values.get("amount")),
            dividend_rate=_to_decimal(values.get("dividend_rate")),
            brokerage=_to_decimal(values.get("brokerage")),
            stt_charges=_to_decimal(values.get("stt_charges")),
            service_tax=_to_decimal(values.get("service_tax")),
            transaction_charges=_to_decimal(values.get("transaction_charges")),
            stamp_duty=_to_decimal(values.get("stamp_duty")),
            other_charges=_to_decimal(values.get("other_charges")),
            tds=_to_decimal(values.get("tds")),
            notes=_str(values.get("notes")),
            transaction_id=_str(values.get("transaction_id")),
            transaction_qualifier=_str(values.get("transaction_qualifier")),
            transfer_date=_to_date(values.get("transfer_date")),
            transfer_price=_to_decimal(values.get("transfer_price")),
            transfer_amount=_to_decimal(values.get("transfer_amount")),
        )
        errors: list[str] = []
        if not row.trade_date:
            errors.append("TRADE/TRANSACTION DATE missing")
        if not row.transaction_type:
            errors.append("TRANSACTION_TYPE missing")
        if row.amount is None and row.units is None:
            errors.append("AMOUNT/UNITS missing")
        if vehicle == "mutual-fund":
            if not row.folio_number and not row.scheme_name and not row.isin:
                errors.append("FOLIO/SCHEME/ISIN missing")
        else:
            if not row.isin and not row.security_name:
                errors.append("ISIN/SECURITY missing")
        if errors:
            row.iserror = 1
            row.errordesc = "; ".join(errors)
        parsed_rows.append(row)

    return VehicleParseResult(
        vehicle_type=vehicle,
        sheet_name=sheet_name,
        headers=canonical,
        rows=parsed_rows,
    )


def row_to_mf_parsed(row: ParsedVehicleRow):
    """Adapt to mutualfund repository MfParsedRow."""
    from backend.vehicle.mf_excel import MfParsedRow

    return MfParsedRow(
        entity_id=row.entity_id,
        investor_name=None,
        mf_name=row.mf_name,
        folio_number=row.folio_number,
        isin=row.isin,
        scheme_name=row.scheme_name or row.security_name,
        trade_date=row.trade_date,
        transaction_type=row.transaction_type,
        dividend_rate=row.dividend_rate,
        amount=row.amount,
        units=row.units,
        price=row.price,
        transfer_date=row.transfer_date,
        transfer_price=row.transfer_price,
        transfer_amount=row.transfer_amount,
        brokerage=row.brokerage,
        stt_charges=row.stt_charges,
        service_tax=row.service_tax,
        transaction_charges=row.transaction_charges,
        stamp_duty=row.stamp_duty,
        other_charges=row.other_charges,
        tds=row.tds,
        notes=row.notes,
        transaction_id=row.transaction_id,
        transaction_qualifier=row.transaction_qualifier,
        currency=None,
        iserror=row.iserror,
        errordesc=row.errordesc,
    )


def row_to_staging_dict(row: ParsedVehicleRow) -> dict[str, Any]:
    name = row.security_name or row.scheme_name or row.mf_name or ""
    folio = row.folio_number or row.isin or ""
    sync = " | ".join(p for p in [name, folio, row.transaction_type] if p)
    return {
        "entity_id": row.entity_id,
        "trade_date": row.trade_date,
        "transaction_type": row.transaction_type,
        "isin": row.isin,
        "security_name": name,
        "units": row.units,
        "price": row.price,
        "amount": row.amount,
        "syncdescription": sync[:500] if sync else None,
        "iserror": row.iserror,
        "errordesc": row.errordesc,
    }
